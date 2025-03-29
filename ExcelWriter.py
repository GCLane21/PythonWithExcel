import openpyxl
import shutil
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from PracticeData.Data1 import sample_data

class ExcelWriter:
    def __init__(self, filename: str):
        # Set up file name and placeholders for workbook and active sheet
        self.filename = filename
        self.wb = None
        self.sheet = None

    def create_new_workbook(self, sheet_name: str = "Sheet 1"):
        # Create a new workbook and assign a title to the default active sheet
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = sheet_name
        print(f"New workbook created with sheet '{sheet_name}'")

    def load_workbook(self):
        # Backup the original file before loading and modifying it
        backup_filename = self.filename.replace(".xlsx", "_backup.xlsx")
        shutil.copyfile(self.filename, backup_filename)

        # Load workbook and set the active sheet
        self.wb = load_workbook(filename=self.filename)
        self.sheet = self.wb.active
        print(f"Workbook '{self.filename}' loaded with backup '{backup_filename}'")

    def create_new_sheet(self, name: str):
        # Add a new sheet to the workbook with a specified name
        self.sheet = self.wb.create_sheet(title=name)
        print(f"New sheet '{name}' created")

    def write_dict(self, data):
        # Write a list of dictionaries as rows into the sheet
        if not data:
            raise IndexError("Data list is empty")

        headers = list(data[0].keys())

        # Write column headers
        for col_num, header in enumerate(headers, 1):
            self.sheet.cell(row=1, column=col_num, value=header)

        # Write row values using dictionary keys
        for row_num, entry in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                self.sheet.cell(row=row_num, column=col_num, value=entry.get(header, None))

        print("Dictionary data written to sheet")

    def write_dataframe(self, dataframe: pd.DataFrame):
        # Skip writing if the DataFrame is empty
        if dataframe.empty:
            print("Empty DataFrame. Nothing was written.")
            return

        # Write DataFrame headers
        for col_num, column in enumerate(dataframe.columns, 1):
            self.sheet.cell(row=1, column=col_num, value=column)

        # Write row values from DataFrame
        for row_num, row in enumerate(dataframe.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                self.sheet.cell(row=row_num, column=col_num, value=value)

        print("DataFrame written to sheet")

    def add_dict(self, new_data):
        # Append a single dictionary to the next available row based on existing headers
        new_row = self.sheet.max_row + 1

        # Extract header names from first row
        headers = [self.sheet.cell(row=1, column=col).value for col in range(1, self.sheet.max_column + 1)]

        # Write values aligned to headers, using None for missing keys
        for col_num, header in enumerate(headers, 1):
            self.sheet.cell(row=new_row, column=col_num, value=new_data.get(header, None))

        print("Dictionary entry added")

    def add_dataframe(self, dataframe: pd.DataFrame):
        # Append a DataFrame starting from the next empty row
        start_row = self.sheet.max_row + 1

        # Write each DataFrame row into the sheet
        for row_num, row in enumerate(dataframe.itertuples(index=False), start_row):
            for col_num, value in enumerate(row, 1):
                self.sheet.cell(row=row_num, column=col_num, value=value)

        print(f"Appended DataFrame starting from row {start_row}")

    def edit_cell(self, cell_reference: str, value):
        # Validate cell reference format before editing
        try:
            col, row = coordinate_from_string(cell_reference)
            col_idx = column_index_from_string(col)
            row = int(row)
            if col_idx < 1 or col_idx > 16384 or row < 1 or row > 1048576:
                raise ValueError(f"Cell reference out of bounds: {cell_reference}")
        except Exception:
            raise ValueError(f"Invalid cell reference: {cell_reference}")
    

        # Assign new value to the specified cell
        self.sheet[cell_reference] = value
        print(f"Updated {cell_reference} to '{value}'")

    def save(self):
        # Save changes to the original file
        self.wb.save(self.filename)
        print(f"Workbook saved as '{self.filename}'")

    def save_to_location(self, folder_path: str, new_filename: str = None):
        # Save the workbook to a different location, optionally with a new filename
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Determine output path
        filename_to_use = new_filename if new_filename else os.path.basename(self.filename)
        full_path = os.path.join(folder_path, filename_to_use)

        # Save the workbook to the new path
        self.wb.save(full_path)
        print(f"Workbook saved to '{full_path}'")


# Example usage of the ExcelWriter class
if __name__ == "__main__":
    writer = ExcelWriter("sample_excel_file.xlsx")

    # Create and populate a new workbook
    writer.create_new_workbook("Sample Data")
    writer.write_dict(sample_data)
    writer.edit_cell("B3", "Updated Name")
    writer.save()

    # Write a DataFrame to a new sheet
    df = pd.DataFrame(sample_data)
    writer.load_workbook()
    writer.create_new_sheet("DataFrame Data")
    writer.write_dataframe(df)
    writer.save()

    # Save the final result to a different location
    writer.save_to_location("./exports", "backup_export.xlsx")